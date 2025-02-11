import requests
import pandas as pd
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

class CryptoTracker:
    def __init__(self):
        self.base_url = "https://api.coingecko.com/api/v3"
        self.excel_file = "crypto_tracker.xlsx"
        
    def fetch_top_50_data(self):
        """Fetch top 50 cryptocurrencies data from CoinGecko"""
        endpoint = f"{self.base_url}/coins/markets"
        params = {
            'vs_currency': 'usd',
            'order': 'market_cap_desc',
            'per_page': 50,
            'page': 1,
            'sparkline': False
        }
        
        try:
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            return response.json()
        except requests.RequestException as e:
            print(f"Error fetching data: {e}")
            return None

    def process_data(self, data):
        """Process the raw data into a pandas DataFrame"""
        if not data:
            return None
        
        df = pd.DataFrame(data)
        df = df[[
            'name', 'symbol', 'current_price', 'market_cap',
            'total_volume', 'price_change_percentage_24h'
        ]]
        df.columns = [
            'Name', 'Symbol', 'Price (USD)', 'Market Cap (USD)',
            '24h Volume (USD)', '24h Change (%)'
        ]
        return df

    def analyze_data(self, df):
        """Perform analysis on the cryptocurrency data"""
        analysis = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'top_5_by_market_cap': df.head(5)[['Name', 'Symbol', 'Market Cap (USD)', 'Price (USD)']].to_dict('records'),
            'average_price': df['Price (USD)'].mean(),
            'median_price': df['Price (USD)'].median(),
            'total_market_cap': df['Market Cap (USD)'].sum(),
            'highest_24h_change': df.nlargest(1, '24h Change (%)')['Name'].iloc[0],
            'highest_24h_change_value': df['24h Change (%)'].max(),
            'lowest_24h_change': df.nsmallest(1, '24h Change (%)')['Name'].iloc[0],
            'lowest_24h_change_value': df['24h Change (%)'].min(),
            'volatile_coins': df.nlargest(3, '24h Change (%)').abs()[['Name', '24h Change (%)']].to_dict('records')
        }
        return analysis

    def update_excel(self, df, analysis):
        """Update Excel file with latest data and analysis"""
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Live Data', index=False)
            
            # Format Live Data sheet
            worksheet = writer.sheets['Live Data']
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Add status indicator and last update time at the top
            status_data = pd.DataFrame([
                ['Status', 'LIVE'],
                ['Last Updated', analysis['timestamp']],
                ['Next Update In', '5 minutes'],
                ['', '']  # Empty row before main data
            ])
            status_data.to_excel(writer, sheet_name='Live Data', startrow=0, index=False)
            
            # Format status indicators
            status_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # Green color
            worksheet['B1'].fill = status_fill
            worksheet['B1'].font = Font(color='FFFFFF', bold=True)
            
            # Format main data headers
            for cell in worksheet[5]:  # Adjusted to account for status rows
                cell.fill = header_fill
                cell.font = header_font
            
            # Create analysis sheet with enhanced metrics
            analysis_data = [
                ['Metric', 'Value'],
                ['Last Updated', analysis['timestamp']],
                ['Average Price (USD)', f"${analysis['average_price']:.2f}"],
                ['Median Price (USD)', f"${analysis['median_price']:.2f}"],
                ['Total Market Cap (USD)', f"${analysis['total_market_cap']:,.2f}"],
                ['Highest 24h Change', f"{analysis['highest_24h_change']} ({analysis['highest_24h_change_value']:.2f}%)"],
                ['Lowest 24h Change', f"{analysis['lowest_24h_change']} ({analysis['lowest_24h_change_value']:.2f}%)"],
                ['', ''],
                ['Top 5 by Market Cap:', '']
            ]
            
            # Add top 5 cryptocurrencies details
            for i, coin in enumerate(analysis['top_5_by_market_cap'], 1):
                analysis_data.append([
                    f"{i}. {coin['Name']} ({coin['Symbol']})",
                    f"Market Cap: ${coin['Market Cap (USD)']:,.2f}"
                ])
            
            analysis_data.append(['', ''])
            analysis_data.append(['Most Volatile Coins (24h):', ''])
            
            for coin in analysis['volatile_coins']:
                analysis_data.append([
                    f"{coin['Name']}", 
                    f"{coin['24h Change (%)']:.2f}%"
                ])
            
            analysis_df = pd.DataFrame(analysis_data)
            analysis_df.to_excel(writer, sheet_name='Analysis', index=False)
            
            # Format Analysis sheet
            worksheet = writer.sheets['Analysis']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

    def run(self, update_interval=300):
        """Run the tracker with specified update interval (in seconds)"""
        print("Starting Crypto Tracker...")
        print(f"Data will update every {update_interval} seconds")
        
        while True:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"\nFetching data at {current_time}")
            
            try:
                # Fetch and process data
                raw_data = self.fetch_top_50_data()
                if raw_data:
                    df = self.process_data(raw_data)
                    analysis = self.analyze_data(df)
                    self.update_excel(df, analysis)
                    print("✓ Excel file updated successfully")
                    print(f"Next update in {update_interval/60:.1f} minutes")
                
                # Wait for the next update
                time.sleep(update_interval)
            except Exception as e:
                print(f"✗ An error occurred: {e}")
                print("Waiting 60 seconds before retrying...")
                time.sleep(60)

if __name__ == "__main__":
    tracker = CryptoTracker()
    # Update every 5 minutes (300 seconds)
    tracker.run(update_interval=300)
    